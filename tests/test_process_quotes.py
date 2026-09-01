from openpyxl import Workbook, load_workbook

from main import process_quotes


def test_process_quotes_end_to_end(tmp_path):
    # Create temporary input workbook
    input_file = tmp_path / "test_input.xlsx"
    output_file = tmp_path / "test_output.xlsx"

    workbook = Workbook()
    sheet = workbook.active

    sheet.append([
        "SKU",
        "Quantity",
        "Unit Cost",
        "Shipping Cost",
        "Markup"
    ])

    # Valid row
    sheet.append([
        "GOOD001",
        500,
        4.25,
        350,
        20
    ])

    # Invalid row
    sheet.append([
        "BAD001",
        0,
        -5,
        -100,
        -10
    ])

    workbook.save(input_file)

    # Run the actual application pipeline
    summary = process_quotes(
        str(input_file),
        str(output_file)
    )

    # Verify summary
    assert summary["successful"] == 1
    assert summary["errors"] == 1

    # Verify output file was created
    assert output_file.exists()

    # Open generated workbook
    result_workbook = load_workbook(output_file)

    assert "Results" in result_workbook.sheetnames
    assert "Errors" in result_workbook.sheetnames
    
    results_sheet = result_workbook["Results"]
    errors_sheet = result_workbook["Errors"]

    assert results_sheet["A2"].value == "GOOD001"
    assert results_sheet["F2"].value == 0.70
    assert results_sheet["G2"].value == 4.95
    assert results_sheet["H2"].value == 5.94

    assert errors_sheet["A2"].value == "BAD001"
    
    error_message = errors_sheet["B2"].value

    assert "Quantity must be a positive integer." in error_message
    assert "Unit cost must be non-negative." in error_message
    assert "Shipping cost must be non-negative." in error_message
    assert "Markup must be between 0 and 100." in error_message