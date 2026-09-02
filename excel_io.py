from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def read_quotes(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    
    rows = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku, quantity, unit_cost, shipping_cost, markup = row
        
        quote = {
            'sku': sku,
            'quantity': quantity,
            'unit_cost': unit_cost,
            'shipping_cost': shipping_cost,
            'markup': markup
        }
        
        rows.append(quote)
    return rows



def write_results(filename, results, errors):
    workbook = Workbook()
    
    results_sheet = workbook.active
    results_sheet.title = "Results"
    
    errors_sheet = None
        
    # Headers
    results_sheet.append([
        'SKU',
        'Quantity',
        'Unit Cost',
        'Shipping Cost',
        'Markup',
        'Shipping per Unit',
        'Landed Cost',
        'Quote Price'
    ])  
        
    # Results
    for result in results:
        results_sheet.append([
            result['sku'],
            result['quantity'],
            result['unit_cost'],
            result['shipping_cost'],
            result['markup'],
            result['shipping_per_unit'],
            result['landed_cost'],
            result['quote_price']
        ])
    
    # Errors
    
    if errors:
        errors_sheet = workbook.create_sheet(title="Errors")
        errors_sheet.append([
            'SKU',
            'Errors'
        ])
    
    for error in errors:
        error_string = '; '.join(error['errors'])
        
        errors_sheet.append([
            error['sku'],
            error_string
        ])
        
    for cell in results_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    results_sheet.freeze_panes = results_sheet['A2']

    if errors_sheet is not None:
        for cell in errors_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        errors_sheet.freeze_panes = errors_sheet['A2']
        errors_sheet.column_dimensions['A'].width = 15
        errors_sheet.column_dimensions['B'].width = 150
    
    for row in range(2, results_sheet.max_row + 1):
        results_sheet.cell(row=row, column=3).number_format = '$0.00'
        results_sheet.cell(row=row, column=4).number_format = '$0.00'
        results_sheet.cell(row=row, column=6).number_format = '$0.00'
        results_sheet.cell(row=row, column=7).number_format = '$0.00'
        results_sheet.cell(row=row, column=8).number_format = '$0.00'
    
    for row in range(2, results_sheet.max_row + 1):
        results_sheet.cell(row=row, column=5).number_format = '0.00"%"'
    
    column_widths = {
        'A': 15,
        'B': 12,
        'C': 14,
        'D': 16,
        'E': 12,
        'F': 18,
        'G': 15,
        'H': 15
    }
    
    for column, width in column_widths.items():
        results_sheet.column_dimensions[column].width = width
        
    if results:
        table_range = f"A1:H{results_sheet.max_row}"
        
        results_table = Table(
            displayName='QuoteResults',
            ref=table_range
        )
        
        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        results_table.tableStyleInfo = table_style
        results_sheet.add_table(results_table)
        
    
    workbook.save(filename)
    
    
